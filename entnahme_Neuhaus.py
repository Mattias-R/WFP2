# importing openpyxl module
import openpyxl

# input excel file path
inputExcelFile =r'C:\Users\matti\PycharmProjects\csvIntoJson\verbrauchsdaten_Neuhaus.xlsx'

# creating or loading an excel workbook
newWorkbook = openpyxl.load_workbook(inputExcelFile)
sheet_obj = newWorkbook.active
print(sheet_obj.max_row)

# Accessing specific sheet of the workbook.
firstWorksheet = newWorkbook["Tabelle1"]

maxLines = sheet_obj.max_row

i = 1
datei = open(r'C:\Users\matti\PycharmProjects\csvIntoJson\sendData.txt', 'a')
datei.write("\r\n")
x = 0
#-----------------------------------------------------------------------
# Damit kann der token für jedes erstellte Gerät eingefügt und durchgearbetiet werden. Die meiste arbeit ist nur noch das Gerät zu erstellen
token = ["DJxouOYQt2tJHbmN9w9p", "DJxouOYQt2tJHbmN9w9p", "CYhJRqWVBDqlZTTp0WVF", "CYhJRqWVBDqlZTTp0WVF","Mehk8cFJWbOLGmdoochq","eIdfmgIQUNPS0xgYsc5j","OPodGd0aw5ls8BINFVlL","CyfvULC9ucWrWq39M2oe","dG9gadBXhgUypVaeZmc0","AX9KDg9XWUwSK7UwMjlC","zyYpoqmma1xNSWOJD3ZA","R5LDP3uzVhKkTXGMpVNC","6m1nP5xyHGVDxZWUNNUI","sdGiYGRJT9Ty5QXPktGn","VM8T8uBLPBlWYt7JOFSo","6aflvlwLXdFbRjnAEcXO","BQpCUVmeJZaqqUaOKFnZ","0jWgAUZMEBdkfg6ssyKd","gcZHRQV0e8o3fzbkGvzk","CODbx6es6yGUe1OzoXJU","gcpQ1vm1KqJ9KOS7djCp","Z5O2gRnzDio94frZgP3L","RrevNDStrumW4SdvTQWu","YFv20ahTGjgWok7D56R3","KiPNM8dtyF1UEbZunTzL","Rr3sqJowOCNeiktpMxLL","K5XnC18tCiIsXE5MgyLc","WTYtM0tfw17qeZWVvsrq","djjmotN4L3CWxevJN7zm","GdrYIeU4VV7wMJ3av5DJ","CJE6xMgEoLZYszzZ9K8Q","PDwgNAyHTV86MHczp6aT","k31v8e7n8DrgxYgQ0Hon","sKrI8NApMaXiYUoX2Drr","neYCdxRgVlHwWiFIlmjf","oM5eR1A0fJmlWGdm3u3K","SuDri8EbCsUuFYsG9PSC","bdKrXThuRzfjt8qmCiQO","tiOeGzp7uKKHUskvAhno","NvFTDgNz10PRYwzBouEh","NGQszqUdmX5YEVcW4ElE","n7VGEdi6ii9rCqMuKvKj","fSluc4SgvtXTAdSna9KF","wBG9xtZWkcxCVN2xVxu5","9lt4kv6C4Wu5eNClmh7j","13rBQQtrLleguBg0D8vd","kjt64aDx0ijdlylLwL5W","1YTjW6wRFbR55UAuySGH","U3iENVlRG3NT0Z6aPWNl","H1QQuMSxXJjR7YTdcndm","W60EEDeHbZcuHmLyQKqJ","7wu6yK5Ss9Zpnkbiia5b","osD8c7DxiyEOziRromL3","TZT5adJgwWP2nKC5LUv7","c8uf2gNi4CeCWluLfvaj","KRNIbPIpfkQWuTFwoVsc","GGumqqME1Dq7QwRHso4P","52uWI2orlaDlQrEZl2aM","QahLClN2fXfxlE4J5YSM","Kpb7C99kPrp07yK1bXcT","yWB6R2MyLAbZIZ98kkxD","C9oBQMBdaSGwmgb8kyLK","zcKTcJ9eXzxEiZCoRoiO","vIcEVcLjuHrk6nNm2UCE","62N4nTZAXOjheToXXKWY","sL1zlFnL0emnwuosNY3r","Jh1Q11ibpkt3xi8lGHeH","kd3F4m7JKzgB9E5HPhk1","FguyXKlAq2zcEGMxYn4s","AXwbfUWD6WbeCHVnLLg9","HwpA3CMr0brwytmuTomT","YnLAi2u4NdUftnvO7pRJ","40dXzQv3UAiF9HTmlmE1","WaW6dSUeDlYU657epCKZ"]
tok = ["DJii10uC1VDxT7224lvw"]
# Ascii -> A bis Z ist 65 bis 90. anfangen tun wir mit B, also 66
table = 67
table2 = 64
bol = 0
bol2 = 0
letzerWert = 0
helper = 1
helper2 = 0
werte = []
#für jedes gerät gibt es eine spalte, deswegen die schleife mit den devices.
#nach jeder iteration wird die nächste Excel spalte herangezogen
for device in token:
   if(table == 91):
      table = 65
      bol = 1
      table2 = table2 + 1
   #diese Schleife ist dafür da, dass alle werte einer spalte bearbeitet werden.
   #WENN DIE ZAHL ÜBER 1000 IST DANN MUSS SIE ALS NA GEWERTET WERDEN, WEIL MEHRERE DATUM SICH EINGESCHLICHEN HABEN
   for i in range(maxLines):

      if(i == 0):
         i = i + 1

      timestampp = firstWorksheet["A"]
      timestampp = timestampp[i]

      if(bol == 1):
         tableName = firstWorksheet["" + chr(table2) + chr(table)]
         tableName = tableName[i]
         print(chr(table2) + chr(table))
      else:
         tableName = firstWorksheet["" + chr(table)]
         tableName = tableName[i]
      #print(chr(table))

      #wenn NA dann x+1 damit man den sprung zurück weiß
      if(tableName.value != "NA" and float(tableName.value) > 3000):
         tableName.value = "NA"

      if(tableName.value == "NA"):
         #print("NA - erstes if")
         x = x+1
         werte.append(timestampp.value)

      else:

         if(x == 0):
            #print(tableName.value)
            #mit dem letzten wert werden die leeren felder berechnet
            letzerWert = tableName.value
            #werte.append(tableName.value)
            datei.write("sudo curl -v -X POST --data \"{\"ts\":")
            datei.write(str(timestampp.value))
            datei.write(",\"values\":{\"value\":")
            datei.write(str(tableName.value))
            datei.write("}}\" localhost:8080/api/v1/")
            datei.write(device)
            datei.write("/telemetry --header \"Content-Type:application/json\"\r\n")
         else:
            innerLoop = x
            helper = x
            x = 0
            help = letzerWert
            for k in range(innerLoop):
               #print(letzerWert)
               #print(tableName.value)
               #print(float(letzerWert)+(float(tableName.value)-float(letzerWert))/(helper+1))
               letzerWert = float(letzerWert)+(float(tableName.value)-float(help))/(helper+1)
               datei.write("sudo curl -v -X POST --data \"{\"ts\":")
               datei.write(str(werte[k]))
               datei.write(",\"values\":{\"value\":")
               round(letzerWert, 3)
               datei.write(str(letzerWert))
               datei.write("}}\" localhost:8080/api/v1/")
               datei.write(device)
               datei.write("/telemetry --header \"Content-Type:application/json\"\r\n")
               #datei.write("sleep 10\r\n")
            datei.write("sudo curl -v -X POST --data \"{\"ts\":")
            datei.write(str(timestampp.value))
            datei.write(",\"values\":{\"value\":")
            datei.write(str(tableName.value))
            datei.write("}}\" localhost:8080/api/v1/")
            datei.write(device)
            datei.write("/telemetry --header \"Content-Type:application/json\"\r\n")
            #datei.write("sleep 10\r\n")
         werte.clear()
   table = table + 1


