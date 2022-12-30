import paho.mqtt.client as paho  # mqtt library
import os
import json
import time
from datetime import datetime

import openpyxl

# input excel file path
inputExcelFile =r'C:\Users\matti\PycharmProjects\csvIntoJson\verbrauchsdaten_Neuhaus.xlsx'

newWorkbook = openpyxl.load_workbook(inputExcelFile)
sheet_obj = newWorkbook.active
print(sheet_obj.max_row)
firstWorksheet = newWorkbook["Tabelle1"]
maxLines = sheet_obj.max_row
datei = open(r'C:\Users\matti\PycharmProjects\csvIntoJson\sendData.txt', 'a')
tok = ["eJ9fFvIdqCoM07HhWVGK"]
tableName = firstWorksheet["C"]


ACCESS_TOKEN = 'Ak32uCMrE8UxeH1kaBs0'  # Token of your device+
broker = "77.237.53.201"  # host name
port = 13883  # data listening port


def on_publish(client, userdata, result):  # create function for callback
    print("data published to thingsboard \n")
    pass

client1 = paho.Client("control1")  # create client object
client1.on_publish = on_publish  # assign function to callback
client1.username_pw_set(ACCESS_TOKEN)  # access token from thingsboard device
#client1.username_pw_set("marbanriedl", "WFP1")
client1.connect(broker, port)  # establish connection
timestampp = firstWorksheet["A"]



for i in range(maxLines):
    if (i == 0):
        continue
    if (tableName[i].value != "NA" and float(tableName[i].value) > 3000):
        tableName[i].value = "NA"
    # skip loop if value is NA
    if (tableName[i].value == "NA"):
        continue
    # skip loop if value is WM
    if (str(tableName[i - 1].value).startswith("WM")):
        continue
    #payload = "{"
    #payload += "\"timestamp\":" + str(timestampp[i].value) + ",";
    #payload += "\"value\":" + tableName[i].value;
    #payload += "}"

    payload = {
            "ts": timestampp[i].value,
            "values": {
                "value": tableName[i].value
            }
    }

    MQTT_MSG = json.dumps(payload)

    ret = client1.publish("v1/devices/me/telemetry", MQTT_MSG)  # topic-v1/devices/me/telemetry
    print("Please check LATEST TELEMETRY field of your device")
    print(payload);
    time.sleep(1)
#newMsg.delta = temperatureArray[temperatureArray.length -1].value