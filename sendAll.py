# importing openpyxl module
import openpyxl
import paho.mqtt.client as paho  # mqtt library
import os
import json
import time
from datetime import datetime

def on_publish(client, userdata, result):  # create function for callback
    print("data published to thingsboard \n")
    pass

# input excel file path
inputExcelFile =r'C:\Users\matti\PycharmProjects\csvIntoJson\verbrauchsdaten_Neuhaus.xlsx'
#inputExcelFile =r'C:\Users\matti\PycharmProjects\csvIntoJson\test.xlsx'
# creating or loading an excel workbook
newWorkbook = openpyxl.load_workbook(inputExcelFile)
sheet_obj = newWorkbook.active
print(sheet_obj.max_row)
# Accessing specific sheet of the workbook.
firstWorksheet = newWorkbook["Tabelle1"]
#firstWorksheet = newWorkbook["test"]
maxLines = sheet_obj.max_row
#-----------------------------------------------------------------------
# Damit kann der token für jedes erstellte Gerät eingefügt und durchgearbetiet werden. Die meiste arbeit ist nur noch das Gerät zu erstellen
token = ["DJxouOYQt2tJHbmN9w9p", "DJxouOYQt2tJHbmN9w9p", "CYhJRqWVBDqlZTTp0WVF", "CYhJRqWVBDqlZTTp0WVF","Mehk8cFJWbOLGmdoochq","eIdfmgIQUNPS0xgYsc5j","OPodGd0aw5ls8BINFVlL","CyfvULC9ucWrWq39M2oe","dG9gadBXhgUypVaeZmc0","AX9KDg9XWUwSK7UwMjlC","zyYpoqmma1xNSWOJD3ZA","R5LDP3uzVhKkTXGMpVNC","6m1nP5xyHGVDxZWUNNUI","sdGiYGRJT9Ty5QXPktGn","VM8T8uBLPBlWYt7JOFSo","6aflvlwLXdFbRjnAEcXO","BQpCUVmeJZaqqUaOKFnZ","0jWgAUZMEBdkfg6ssyKd","gcZHRQV0e8o3fzbkGvzk","CODbx6es6yGUe1OzoXJU","gcpQ1vm1KqJ9KOS7djCp","Z5O2gRnzDio94frZgP3L","RrevNDStrumW4SdvTQWu","YFv20ahTGjgWok7D56R3","KiPNM8dtyF1UEbZunTzL","Rr3sqJowOCNeiktpMxLL","K5XnC18tCiIsXE5MgyLc","WTYtM0tfw17qeZWVvsrq","djjmotN4L3CWxevJN7zm","GdrYIeU4VV7wMJ3av5DJ","CJE6xMgEoLZYszzZ9K8Q","PDwgNAyHTV86MHczp6aT","k31v8e7n8DrgxYgQ0Hon","sKrI8NApMaXiYUoX2Drr","neYCdxRgVlHwWiFIlmjf","oM5eR1A0fJmlWGdm3u3K","SuDri8EbCsUuFYsG9PSC","bdKrXThuRzfjt8qmCiQO","tiOeGzp7uKKHUskvAhno","NvFTDgNz10PRYwzBouEh","NGQszqUdmX5YEVcW4ElE","n7VGEdi6ii9rCqMuKvKj","fSluc4SgvtXTAdSna9KF","wBG9xtZWkcxCVN2xVxu5","9lt4kv6C4Wu5eNClmh7j","13rBQQtrLleguBg0D8vd","kjt64aDx0ijdlylLwL5W","1YTjW6wRFbR55UAuySGH","U3iENVlRG3NT0Z6aPWNl","H1QQuMSxXJjR7YTdcndm","W60EEDeHbZcuHmLyQKqJ","7wu6yK5Ss9Zpnkbiia5b","osD8c7DxiyEOziRromL3","TZT5adJgwWP2nKC5LUv7","c8uf2gNi4CeCWluLfvaj","KRNIbPIpfkQWuTFwoVsc","GGumqqME1Dq7QwRHso4P","52uWI2orlaDlQrEZl2aM","QahLClN2fXfxlE4J5YSM","Kpb7C99kPrp07yK1bXcT","yWB6R2MyLAbZIZ98kkxD","C9oBQMBdaSGwmgb8kyLK","zcKTcJ9eXzxEiZCoRoiO","vIcEVcLjuHrk6nNm2UCE","62N4nTZAXOjheToXXKWY","sL1zlFnL0emnwuosNY3r","Jh1Q11ibpkt3xi8lGHeH","kd3F4m7JKzgB9E5HPhk1","FguyXKlAq2zcEGMxYn4s","AXwbfUWD6WbeCHVnLLg9","HwpA3CMr0brwytmuTomT","YnLAi2u4NdUftnvO7pRJ","40dXzQv3UAiF9HTmlmE1","WaW6dSUeDlYU657epCKZ"]
tok = ["JMTYplSs6rzSwf53yfh5"]
# Ascii -> A bis Z ist 65 bis 90. anfangen tun wir mit B, also 66
table = 67
#table = 76
#table = 66
table2 = 64
bol = 0
# ----------------------------------------------------------------- #
broker = "77.237.53.201"  # host name
port = 13883  # data listening port
timestampp = firstWorksheet["A"]
#timestampp = firstWorksheet["J"]

for device in tok:

   ACCESS_TOKEN = device
   client1 = paho.Client("control1")  # create client object
   client1.on_publish = on_publish  # assign function to callback
   client1.username_pw_set(ACCESS_TOKEN)  # access token from thingsboard device
   client1.connect(broker, port)  # establish connection

   if(table == 91):
      table = 65
      bol = 1
      table2 = table2 + 1

   for i in range(maxLines):

      if(i == 0):
         i = i + 1

      if(bol == 1):
         tableName = firstWorksheet["" + chr(table2) + chr(table)]
         print(chr(table2) + chr(table))
      else:
         tableName = firstWorksheet["" + chr(table)]
      if (tableName[i].value != "NA" and float(tableName[i].value) > 3000):
          tableName[i].value = "NA"
      # skip loop if value is NA
      if (tableName[i].value == "NA"):
          continue
      # skip loop if value is WM
      if (str(tableName[i - 1].value).startswith("WM")):
          continue

      print(tableName[i].value)
      payload = {
          "ts": timestampp[i].value,
          "values": {
              "value": tableName[i].value
          }
      }

      MQTT_MSG = json.dumps(payload)

      ret = client1.publish("v1/devices/me/telemetry", MQTT_MSG)  # topic-v1/devices/me/telemetry
      print("Please check LATEST TELEMETRY field of your device")
      print(payload);
      time.sleep(1)

   client1.disconnect()

   table = table + 1


