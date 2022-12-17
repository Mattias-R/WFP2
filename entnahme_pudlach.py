# importing openpyxl module
import openpyxl

# input excel file path
inputExcelFile =r'C:\Users\matti\PycharmProjects\csvIntoJson\verbrauchsdaten_Neuhaus.xlsx'
# creating or loading an excel workbook
newWorkbook = openpyxl.load_workbook(inputExcelFile)
sheet_obj = newWorkbook.active
print(sheet_obj.max_row)

# Accessing specific sheet of the workbook.
firstWorksheet = newWorkbook["Tabelle1"]

maxLines = sheet_obj.max_row

i = 1
datei = open(r'C:\Users\matti\PycharmProjects\csvIntoJson\sendData.txt', 'a')
datei.write("\r\n")
x = 0
#-----------------------------------------------------------------------
# Damit kann der token für jedes erstellte Gerät eingefügt und durchgearbetiet werden. Die meiste arbeit ist nur noch das Gerät zu erstellen
token = ["test1", "test2", "test3", "test4","test5","test6","test7","test8","test9","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test8","test9","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7","test7"]
# Ascii -> A bis Z ist 65 bis 90. anfangen tun wir mit B, also 66
table = 67
table2 = 64
bol = 0
bol2 = 0
letzerWert = 0
helper = 1
helper2 = 0
werte = []
#für jedes gerät gibt es eine spalte, deswegen die schleife mit den devices.
#nach jeder iteration wird die nächste Excel spalte herangezogen
for device in token:
   if(table == 91):
      table = 65
      bol = 1
      table2 = table2 + 1
   #diese Schleife ist dafür da, dass alle werte einer spalte bearbeitet werden.
   #WENN DIE ZAHL ÜBER 1000 IST DANN MUSS SIE ALS NA GEWERTET WERDEN, WEIL MEHRERE DATUM SICH EINGESCHLICHEN HABEN
   for i in range(maxLines):

      if(i == 0):
         i = i + 1

      timestampp = firstWorksheet["A"]
      timestampp = timestampp[i]

      if(bol == 1):
         tableName = firstWorksheet["" + chr(table2) + chr(table)]
         tableName = tableName[i]
         print(chr(table2) + chr(table))
      else:
         tableName = firstWorksheet["" + chr(table)]
         tableName = tableName[i]
      #print(chr(table))

      #wenn NA dann x+1 damit man den sprung zurück weiß
      if(tableName.value != "NA" and float(tableName.value) > 3000):
         tableName.value = "NA"

      if(tableName.value == "NA"):
         #print("NA - erstes if")
         x = x+1
         werte.append(timestampp.value)

      else:

         if(x == 0):
            #print(tableName.value)
            #mit dem letzten wert werden die leeren felder berechnet
            letzerWert = tableName.value
            #werte.append(tableName.value)
            datei.write("sudo curl -v -X POST --data \"{\"ts\":")
            datei.write(str(timestampp.value))
            datei.write(",\"values\":{\"value\":")
            datei.write(str(tableName.value))
            datei.write("}}\" localhost:8080/api/v1/")
            datei.write(device)
            datei.write("/telemetry --header \"Content-Type:application/json\"\r\n")
         else:
            innerLoop = x
            helper = x
            x = 0
            help = letzerWert
            for k in range(innerLoop):
               #print(letzerWert)
               #print(tableName.value)
               #print(float(letzerWert)+(float(tableName.value)-float(letzerWert))/(helper+1))
               letzerWert = float(letzerWert)+(float(tableName.value)-float(help))/(helper+1)
               datei.write("sudo curl -v -X POST --data \"{\"ts\":")
               datei.write(str(werte[k]))
               datei.write(",\"values\":{\"value\":")
               round(letzerWert, 3)
               datei.write(str(letzerWert))
               datei.write("}}\" localhost:8080/api/v1/")
               datei.write(device)
               datei.write("/telemetry --header \"Content-Type:application/json\"\r\n")
            datei.write("sudo curl -v -X POST --data \"{\"ts\":")
            datei.write(str(timestampp.value))
            datei.write(",\"values\":{\"value\":")
            datei.write(str(tableName.value))
            datei.write("}}\" localhost:8080/api/v1/")
            datei.write(device)
            datei.write("/telemetry --header \"Content-Type:application/json\"\r\n")
         werte.clear()
   table = table + 1


